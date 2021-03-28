#!python3

# multiplicationTable.py - User gives a number 1-10(to command line), we create a multiplication table of that in .xlsx file.

import openpyxl, os, sys
from openpyxl.styles import Font

if len(sys.argv) > 1:
    #Get number from command line
    number = int(''.join(sys.argv[1:]))
else:
    print('Please give a number to command line as argument')

fontObj = Font(bold=True)   # Bolded font to "table" numbers
wb = openpyxl.Workbook()
sheet = wb.active

# First we create a "table" of numbers depending what number user gave us.
for i in range(number):
    if(i == 0):
        for j in range(number):
            sheet.cell(row=1, column=j + 2).value = j + 1
            sheet.cell(row=1, column=j + 2).font = fontObj
    sheet.cell(row=i + 2, column=1).value = i + 1
    sheet.cell(row=i + 2, column=1).font = fontObj

# Let's make the multiplications.

for i in range(number):
    for j in range(number):
        value1 = i + 1 # row value
        value2 = j + 1 # column value
        sheet.cell(row=value1 + 1, column=value2 + 1).value = value1 * value2


wb.save('/Users/janitiainen/Documents/Koodia/Automate boring stuff/13. excel_worksheets/multiplicationTable.xlsx')
